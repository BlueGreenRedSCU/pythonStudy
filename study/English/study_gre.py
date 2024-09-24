#!/usr/bin/env python

# -- coding: utf-8 --
# @Time : 2023/5/15 8:47 PM
# @Author : ritong.lan
# @File : study_gre.py

import sys
import uuid
import requests
import hashlib
import time
from imp import reload
import xlrd
import xlwt
import datetime
from openpyxl import Workbook, load_workbook
from collections import deque


CSV_FILE_PATH = "/Users/ritong.lan/Desktop/Words.xlsx"
YOUDAO_URL = 'https://openapi.youdao.com/api'
APP_KEY = '6f51420c1f7ed913'
APP_SECRET = 'Sh21LAZ1rAnAUSvpdJ7SEnZ1Tt3gVcpY'


class Translator:
    def __init__(self):
        pass

    @staticmethod
    def encrypt(signStr):
        hash_algorithm = hashlib.sha256()
        hash_algorithm.update(signStr.encode('utf-8'))
        return hash_algorithm.hexdigest()

    @staticmethod
    def truncate(q):
        if q is None:
            return None
        size = len(q)
        return q if size <= 20 else q[0:10] + str(size) + q[size - 10:size]

    @staticmethod
    def do_request(data):
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        return requests.post(YOUDAO_URL, data=data, headers=headers)

    def get_translation(self, word):
        q = word
        data = {}
        data['from'] = 'en'
        data['to'] = 'cn'
        data['signType'] = 'v3'
        curTime = str(int(time.time()))
        data['curtime'] = curTime
        salt = str(uuid.uuid1())
        signStr = APP_KEY + self.truncate(q) + salt + curTime + APP_SECRET
        sign = self.encrypt(signStr)
        data['appKey'] = APP_KEY
        data['q'] = q
        data['salt'] = salt
        data['sign'] = sign
        # data['vocabId'] = "您的用户词表ID"
        response = self.do_request(data)
        # print(response.json())
        translations = response.json()["basic"]["explains"]
        # explain as a list
        return translations


def write_record_to_excel(workbook, start_time_object, learned_cnt, success_cnt):
    print(workbook.sheetnames)
    cur_worksheet = workbook[workbook.sheetnames[1]]

    target_row = cur_worksheet.max_row
    end_time_object = datetime.datetime.now()
    start_time = start_time_object.strftime("%H:%M:%S")
    end_time = end_time_object.strftime("%H:%M:%S")
    date = start_time_object.strftime("%Y-%m-%d")

    cur_mock = [date, start_time, end_time, learned_cnt, success_cnt]
    for i in range(len(cur_mock)):
        cur_worksheet.cell(target_row+1, i+1).value = cur_mock[i]
    workbook.save(CSV_FILE_PATH)


def study_failed_words_again(Q):
    while Q:
        word, translation = Q.popleft()
        print("Word:    " + word)
        keyboard_input = input("Do you know? 1 for yes, 2 for no\n")
        while keyboard_input not in ["1", "2"]:
            keyboard_input = input("Do you know? 1 for yes, 2 for no\n")
        if keyboard_input == "1":
            print("Congratulation!\n")
        else:
            for explain in translation:
                print(explain)
            Q.append([word, translation])
    print("You have finished all the words!")


def start_learning(row):
    translator = Translator()
    # print(translator.get_translation("euphonious"))
    row_number = row - 1
    start_time_object = datetime.datetime.now()
    learned_cnt, success_cnt = 0, 0

    workbook = load_workbook(CSV_FILE_PATH)
    worksheet = workbook[workbook.sheetnames[0]]
    worklist = list(worksheet.rows)
    wordsNeedSecondStudy = deque([])

    while True:
        try:
            row_data = worklist[row_number]
            if not row_data[0].value:
                break
            if not row_data[5].value:
                row_data[5].value = 0
            elif row_data[5].value == 1:
                row_number += 1
                continue
            if not row_data[4].value:
                translation = translator.get_translation(row_data[0].value)
                row_data[4].value = " | ".join(translation)
            else:
                translation = row_data[4].value.split(" | ")

            print("Word:   " + row_data[0].value)

            # 1: know; 2: don't know; 3: don't know previous word; 4: stop learning
            keyboard_input = input("Do you know?\n")
            while keyboard_input not in ["1", "2", "3", "4"]:
                keyboard_input = input("Do you know?\n")
            row_data[2].value = row_data[2].value + 1

            if keyboard_input == "1":
                row_data[3].value = row_data[3].value + 1
                success_cnt += 1
                row_number += 1
            elif keyboard_input == "2":
                wordsNeedSecondStudy.append([row_data[0].value, translation])
                row_number += 1
            elif keyboard_input == "3":
                before_row_data = worklist[row_number-1]
                before_row_data[3].value = before_row_data[3].value - 1
                wordsNeedSecondStudy.append([before_row_data[0].value, before_row_data[4].value.split(" | ")])
                continue
            elif keyboard_input == "4":
                break
            else:
                row_data[5].value = 1
                row_number += 1
            for explain in translation:
                print(explain)
            print("\n")
            learned_cnt += 1
        except Exception as e:
            print(e)
            print("\nError occurs, will retry!\n")

    write_record_to_excel(workbook, start_time_object, learned_cnt, success_cnt)
    workbook.save(CSV_FILE_PATH)
    print("study is recorded.\n")

    study_failed_words_again(wordsNeedSecondStudy)


if __name__ == "__main__":
    row = input("Which line to start\n")
    # times = input("Below which times?\n")
    start_learning(int(row))

